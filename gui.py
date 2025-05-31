import asyncio
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import logging
from queue import Queue, Empty
from PIL import Image, ImageTk, ImageDraw, ImageFont
import shutil
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

from webdownloader import async_download_manager
import config

class ImageDownloaderGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Cataloginator")
        self.root.minsize(400, 400)  # Set minimum size to prevent collapse
        self.root.geometry("600x600")  # Initial size, can be resized
        # Set custom icon for main window
        try:
            icon_path = os.path.join(os.path.dirname(__file__), "cataloginator.ico")
            self.root.iconbitmap(icon_path)
        except Exception as e:
            logging.warning(f"Failed to set main window icon: {e}")

        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(pady=10, fill="both", expand=True)

        # Download tab
        self.download_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.download_frame, text="Download")
        self.setup_download_tab()

        # Catalog tab
        self.catalog_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.catalog_frame, text="Catalog")
        self.setup_catalog_tab()

        # Footer label
        self.footer_label = tk.Label(self.root, text="made by vP v0.5.1", font=("Arial", 8), fg="gray")
        self.footer_label.pack(side="bottom", pady=5)

        # Queue for progress updates
        self.progress_queue = Queue()

    def setup_download_tab(self):
        # Excel file selection
        self.label_excel = tk.Label(self.download_frame, text="Excel File Destination:")
        self.label_excel.pack(pady=10)
        self.entry_excel = tk.Entry(self.download_frame, width=40)
        self.entry_excel.pack()
        self.button_browse_excel = tk.Button(self.download_frame, text="Browse", command=self.browse_excel)
        self.button_browse_excel.pack(pady=10)

        # Save folder selection
        self.label_folder = tk.Label(self.download_frame, text="Save Folder Destination:")
        self.label_folder.pack(pady=10)
        self.entry_folder = tk.Entry(self.download_frame, width=40)
        self.entry_folder.pack()
        self.button_browse_folder = tk.Button(self.download_frame, text="Browse", command=self.browse_folder)
        self.button_browse_folder.pack(pady=10)

        # Download button
        self.button_download = tk.Button(self.download_frame, text="Download", command=self.start_download)
        self.button_download.pack(pady=20)

        # Progress bar
        self.progress_bar = ttk.Progressbar(self.download_frame, orient="horizontal", length=330, mode="determinate")
        self.progress_bar.pack(pady=10)
        self.progress_label = tk.Label(self.download_frame, text="")
        self.progress_label.pack()

    def setup_catalog_tab(self):
        # Save folder selection
        self.label_catalog_folder = tk.Label(self.catalog_frame, text="Catalog Folder Destination:")
        self.label_catalog_folder.pack(pady=10)
        self.entry_catalog_folder = tk.Entry(self.catalog_frame, width=40)
        self.entry_catalog_folder.pack()
        self.button_browse_catalog_folder = tk.Button(self.catalog_frame, text="Browse", command=self.browse_catalog_folder)
        self.button_browse_catalog_folder.pack(pady=10)

        # Begin button
        self.button_begin = tk.Button(self.catalog_frame, text="Begin", command=self.start_cataloging)
        self.button_begin.pack(pady=20, side="bottom")

    def browse_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.entry_excel.delete(0, tk.END)
            self.entry_excel.insert(0, file_path)

    def browse_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.entry_folder.delete(0, tk.END)
            self.entry_folder.insert(0, folder_path)

    def browse_catalog_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.entry_catalog_folder.delete(0, tk.END)
            self.entry_catalog_folder.insert(0, folder_path)

    def start_download(self):
        excel_file = self.entry_excel.get()
        save_folder = self.entry_folder.get()

        if not excel_file or not save_folder:
            messagebox.showerror("Error", "Please select both Excel file and save folder.")
            return

        # Disable button and reset progress
        self.button_download.config(state="disabled")
        self.progress_bar["value"] = 0
        self.progress_label.config(text="Starting download...")

        # Start download in a separate thread
        threading.Thread(
            target=self.run_download, args=(excel_file, save_folder), daemon=True
        ).start()
        # Start polling for progress
        self.update_progress()

    def run_download(self, excel_file, save_folder):
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            success, error = loop.run_until_complete(
                async_download_manager(excel_file, save_folder, self.progress_queue)
            )
            loop.close()
            self.root.after(0, self.show_result, success, error, save_folder)
        except Exception as e:
            self.root.after(0, self.show_result, False, str(e), save_folder)

    def update_progress(self):
        try:
            while True:
                current, total = self.progress_queue.get_nowait()
                percentage = (current / total) * 100
                self.progress_bar["value"] = percentage
                self.progress_label.config(text=f"Processed {current}/{total} rows")
        except Empty:
            pass
        if self.button_download["state"] == "disabled":
            self.root.after(100, self.update_progress)

    def show_result(self, success, error, save_folder):
        self.button_download.config(state="normal")
        self.progress_label.config(text="")
        self.progress_bar["value"] = 0

        if success:
            image_count = len(
                [f for f in os.listdir(save_folder) if os.path.isfile(os.path.join(save_folder, f))]
            )
            messagebox.showinfo(
                "Success", f"Download completed successfully!\n{image_count} images downloaded."
            )
            logging.debug("Script finished")
        else:
            messagebox.showerror("Error", f"An error occurred: {error}")

    def start_cataloging(self):
        catalog_folder = self.entry_catalog_folder.get()
        if not catalog_folder:
            messagebox.showerror("Error", "Please select a catalog folder.")
            return

        # Create processed and hold folders
        processed_folder = Path("./") / "processed"
        hold_folder = Path("./") / "hold"
        processed_folder.mkdir(exist_ok=True)
        hold_folder.mkdir(exist_ok=True)

        # Get list of images
        image_extensions = ('.jpg', '.jpeg', '.png')
        images = [
            f for f in os.listdir(catalog_folder)
            if os.path.isfile(os.path.join(catalog_folder, f)) and f.lower().endswith(image_extensions)
        ]

        if not images:
            messagebox.showinfo("Info", "No images found in the selected folder.")
            return

        # Initialize Excel report
        self.initialize_excel_report(catalog_folder)

        # Open cataloging window
        self.open_cataloging_window(catalog_folder, images, processed_folder, hold_folder)

    def initialize_excel_report(self, catalog_folder):
        self.excel_path = Path("./") / "catalog_report.xlsx"
        if self.excel_path.exists():
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            logging.debug(f"Loaded existing Excel report at {self.excel_path}")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Catalog Report"
            headers = [
                "", "", "", "", "BWU",
                "Switched OFF", "Screen/SAS", "Header not working", "Low visibility in header",
                "Shelf light", "Adjust shelves", "Top shelf", "Legal issue",
                "Visible content in header", "Short vertical insert", "Shelf light on comp", "Physical damage",
                "Header broken", "BWU not closing", "Broken flap", "Missing shelf",
                "Shelf strip base", "Shelf-strip insert", "Гнушка", "No POSM", "Client price tag over shelfstrip",
                "Header possible to install", "No content in Header", "EMPTY 1", "EMPTY 2", "EMPTY 3",
                "EMPTY 4", "EMPTY 5", "Comment"
            ]
            for col, header in enumerate(headers, 1):
                ws[f"{get_column_letter(col)}1"] = header

            # Set custom column widths
            column_widths = {
                'A': 20,  # bwu
                'B': 20,  # region
                'C': 20,  # Outlet number
                'D': 20,  # Scene id
                'E': 20,  # BWU type
                'F': 25,  # Switched OFF
                'G': 25,  # Screen/SAS
                'H': 25,  # Header not working
                'I': 25,  # Low visibility in header
                'J': 25,  # Shelf light
                'K': 25,  # Adjust shelves
                'L': 25,  # Top shelf
                'M': 25,  # Legal issue
                'N': 25,  # Visible content in header
                'O': 25,  # Short vertical insert
                'P': 25,  # Shelf light on comp
                'Q': 25,  # Physical damage
                'R': 25,  # Header broken
                'S': 25,  # BWU not closing
                'T': 25,  # Broken flap
                'U': 25,  # Missing shelf
                'V': 25,  # Shelf strip base
                'W': 25,  # Shelf-strip insert
                'X': 25,  # Гнушка
                'Y': 25,  # No POSM
                'Z': 30,  # Client price tag over shelfstrip
                'AA': 25, # Header possible to install
                'AB': 25, # No content in Header
                'AC': 15, # EMPTY 1
                'AD': 15, # EMPTY 2
                'AE': 15, # EMPTY 3
                'AF': 15, # EMPTY 4
                'AG': 15, # EMPTY 5
                'AH': 50  # Comment
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Set custom row height for all rows
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 20

            wb.save(self.excel_path)
            logging.debug(f"Initialized new Excel report at {self.excel_path}")

    def open_cataloging_window(self, catalog_folder, images, processed_folder, hold_folder):
        catalog_window = tk.Toplevel(self.root)
        catalog_window.title("Catalog Images")
        # Set custom icon for cataloging window
        try:
            icon_path = os.path.join(os.path.dirname(__file__), "cataloginator.png")
            catalog_window.iconbitmap(icon_path)
        except Exception as e:
            logging.warning(f"Failed to set cataloging window icon: {e}")
        # Maximize window using zoomed, then geometry
        # Set window size to 80% of screen dimensions
        screen_width = catalog_window.winfo_screenwidth()
        screen_height = catalog_window.winfo_screenheight()
        window_width = int(screen_width * 0.8)
        window_height = int(screen_height * 0.8)
        catalog_window.geometry(f"{window_width}x{window_height}+{int(screen_width * 0.1)}+{int(screen_height * 0.1)}")
        catalog_window.minsize(800, 600)  # Ensure minimum size for usability
        # Exit maximized window with Escape key
        catalog_window.bind('<Escape>', lambda e: catalog_window.destroy())

        self.current_image_index = 0
        self.is_zoomed = False

        # Main frame
        main_frame = ttk.Frame(catalog_window)
        main_frame.pack(fill="both", expand=True)

        # Image display on left
        image_frame = ttk.Frame(main_frame)
        image_frame.pack(side="left", padx=20, pady=20)
        self.image_label = tk.Label(image_frame)
        self.image_label.pack(side="top")
        self.image_label.bind("<Button-1>", lambda e: self.toggle_zoom(e, catalog_folder, images, catalog_window))
        self.filename_label = tk.Label(image_frame, text="", font=("arial.ttf", 12))
        self.filename_label.pack(side="top", padx=20, pady=5)

        # Center frame for OK, Hold, Submit buttons
        center_frame = ttk.Frame(main_frame, width=150)  # Fixed width for buttons
        center_frame.pack(side="left", fill="y")
        center_frame.pack_propagate(False)  # Prevent frame from shrinking
        # Inner frame to center buttons vertically
        button_container = ttk.Frame(center_frame)
        button_container.pack(expand=True)
        # OK, Hold, and Submit buttons (in center, vertical layout)
        ok_button = tk.Button(
            button_container, text="OK", bg="green", fg="white", width=14, font=("arial.ttf", 14),
            command=lambda: self.process_image(
                catalog_folder, images, processed_folder, hold_folder, catalog_window, "ok"
            )
        )
        ok_button.pack(pady=10)
        hold_button = tk.Button(
            button_container, text="Hold", bg="yellow", fg="black", width=14, font=("arial.ttf", 14),
            command=lambda: self.process_image(
                catalog_folder, images, processed_folder, hold_folder, catalog_window, "hold"
            )
        )
        hold_button.pack(pady=10)
        submit_button = tk.Button(
            button_container, text="Submit", bg="red", fg="white", width=14, font=("arial.ttf", 14),
            command=lambda: self.process_image(
                catalog_folder, images, processed_folder, hold_folder, catalog_window, "processed"
            )
        )
        submit_button.pack(pady=10)

        # Right frame with scrollable canvas for BWU types, defects, and comments
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side="left", padx=10, pady=10, fill="both", expand=True)  # Expand to fill available space
        canvas = tk.Canvas(right_frame)
        canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=canvas.yview)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar = ttk.Scrollbar(right_frame, orient="horizontal", command=canvas.xview)
        h_scrollbar.pack(side="bottom", fill="x")
        canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        scrollable_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # Update scroll region when frame size changes
        def update_scroll_region(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        scrollable_frame.bind("<Configure>", update_scroll_region)

        # BWU Type (top, mutually exclusive, two rows of 5)
        bwu_frame = ttk.Frame(scrollable_frame)
        bwu_frame.pack(side="top", anchor="w", padx=5, pady=5)
        bwu_label = tk.Label(bwu_frame, text="BWU Type:", font=("arial.ttf", 12))
        bwu_label.pack(side="top", anchor="w")
        self.bwu_var = tk.StringVar(value="")
        bwu_types = [
            "PRO\n", "X\n", "Mini\n", "X\nFLAP", "Mini\nFLAP", "A2\nPr 12/15",
            "SS\nFlaps", "Door\nSlim", "Door\nOval", "Other\n"
        ]
        # Top row (first 5 buttons)
        bwu_top_row = ttk.Frame(bwu_frame)
        bwu_top_row.pack(side="top", fill="x")
        # Bottom row (last 5 buttons)
        bwu_bottom_row = ttk.Frame(bwu_frame)
        bwu_bottom_row.pack(side="top", fill="x")
        self.bwu_buttons = {}

        def select_bwu(bwu_type):
            self.bwu_var.set(bwu_type)
            for btn_type, btn in self.bwu_buttons.items():
                btn.config(
                    bg="red" if btn_type == bwu_type else "gray",
                    activebackground="red" if btn_type == bwu_type else "gray"
                )

        for i, bwu_type in enumerate(bwu_types):
            target_row = bwu_top_row if i < 5 else bwu_bottom_row
            button = tk.Button(
                target_row,
                text=bwu_type,
                bg="gray",
                fg="white",
                activebackground="gray",
                activeforeground="white",
                font=("arial.ttf", 10),
                width=10,
                height=2,
                anchor="center",
                relief="raised",
                command=lambda t=bwu_type: select_bwu(t)
            )
            button.pack(side="left", padx=2, pady=5)
            self.bwu_buttons[bwu_type] = button

        # Detected Defects (middle, non-mutually exclusive)
        defects_frame = ttk.Frame(scrollable_frame)
        defects_frame.pack(pady=10, anchor="w", padx=5)
        defects_label = tk.Label(defects_frame, text="Detected defects:", font=("arial.ttf", 12))
        defects_label.pack(anchor="w")
        self.defect_vars = {}
        self.defect_buttons = {}
        defect_rows = [
            ["Switched\nOFF", "Screen/\nSAS"],
            ["Header\nnot\nworking", "Low\nvisibility\nin header", "Shelf\nlight", "Adjust\nshelves", "Top\nshelf"],
            ["Legal\nissue", "Visible\ncontent\nin header", "Short\nvertical\ninsert", "Shelf light\non comp"],
            ["Physical\ndamage", "Header\nbroken", "BWU not\nclosing", "Broken\nflap", "Missing\nshelf"],
            ["Shelf\nstrip\nbase", "Shelf-strip\ninsert", "Гнушка", "No POSM", "Client price\ntag over\nshelfstrip"],
            ["Header\npossible\nto install", "No content\nin header"],
            ["EMPTY 1", "EMPTY 2", "EMPTY 3", "EMPTY 4", "EMPTY 5"]
        ]

        def toggle_defect(defect, var):
            var.set(not var.get())
            button = self.defect_buttons[defect]
            button.config(
                bg="red" if var.get() else "gray",
                activebackground="red" if var.get() else "gray"
            )

        for row_defects in defect_rows:
            row_frame = ttk.Frame(defects_frame)
            row_frame.pack(fill="x", pady=2)
            for defect in row_defects:
                var = tk.BooleanVar(value=False)
                self.defect_vars[defect] = var
                button = tk.Button(
                    row_frame,
                    text=defect,
                    bg="gray",
                    fg="white",
                    activebackground="gray",
                    activeforeground="white",
                    font=("arial.ttf", 10),
                    width=10,
                    height=3,
                    anchor="center",
                    relief="raised",
                    command=lambda d=defect, v=var: toggle_defect(d, v)
                )
                button.pack(side="left", padx=5, pady=5)
                self.defect_buttons[defect] = button

        # Comments field (within scrollable frame)
        comments_frame = ttk.Frame(scrollable_frame)
        comments_frame.pack(pady=10, anchor="w", padx=5, fill="x")
        comments_label = tk.Label(comments_frame, text="Comments:", font=("arial.ttf", 12))
        comments_label.pack(side="left")
        self.comments_entry = tk.Entry(comments_frame, width=60, font=("arial.ttf", 12), bg="white")
        self.comments_entry.pack(side="left", padx=5, fill="x", expand=True)

        # Load first image and update window
        self.load_image(catalog_folder, images, catalog_window)
        catalog_window.update()

    def load_image(self, catalog_folder, images, catalog_window):
        if self.current_image_index >= len(images):
            # Clear image and show message
            self.image_label.config(image=None)
            self.image_label.config(text="No more images to catalog!", font=("arial.ttf", 24))
            self.filename_label.config(text="")
            return

        image_path = os.path.join(catalog_folder, images[self.current_image_index])
        try:
            # Load and resize image to default thumbnail size
            img = Image.open(image_path)
            max_size = (800, 600)
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self.image_label.config(image=photo, text="")
            self.image_label.image = photo  # Keep reference
            # Update file name label
            self.filename_label.config(text=images[self.current_image_index])
            # Reset zoom state
            self.zoom_level = 0
            # Reset BWU, defect states, and comments
            self.bwu_var.set("")
            for bwu_type, button in getattr(self, 'bwu_buttons', {}).items():
                button.config(bg="gray", activebackground="gray")
            for var in self.defect_vars.values():
                var.set(False)
            for defect, button in getattr(self, 'defect_buttons', {}).items():
                button.config(bg="gray", activebackground="gray")
            if hasattr(self, 'comments_entry'):
                self.comments_entry.delete(0, tk.END)
        except Exception as e:
            logging.error(f"Error loading image {image_path}: {e}")
            self.image_label.config(text="Error loading image", font=("Arial", 20))
            self.filename_label.config(text="")

    def process_image(self, catalog_folder, images, processed_folder, hold_folder, catalog_window, action):
        if self.current_image_index >= len(images):
            return

        current_image = images[self.current_image_index]
        source_path = os.path.join(catalog_folder, current_image)
        ok_folder = Path("./") / "ok"
        if action == "ok":
            ok_folder.mkdir(exist_ok=True)
            dest_folder = ok_folder
        else:
            dest_folder = processed_folder if action == "processed" else hold_folder
        dest_path = os.path.join(dest_folder, current_image)

        try:
            # Save to Excel and add defects to image only on Submit
            if action == "processed":
                self.save_to_excel(current_image)
                self.draw_defects_on_image(source_path)
            # Copy for OK, move for others
            if action == "ok":
                shutil.move(source_path, dest_path)
                logging.debug(f"Copied {current_image} to ok folder")
            else:
                shutil.move(source_path, dest_path)
                logging.debug(f"Moved {current_image} to {action} folder")
        except Exception as e:
            logging.error(f"Error processing {current_image} to {action} folder: {e}")
            messagebox.showerror("Error", f"Failed to process {current_image}: {e}")
            return

        # Move to next image
        self.current_image_index += 1
        self.load_image(catalog_folder, images, catalog_window)

    def save_to_excel(self, image_name):
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            row = ws.max_row + 1

            # Parse filename (assuming format: bwu.region.outlet.scene.jpg)
            parts = image_name.rsplit('.', 4)  # Split on last 4 dots
            if len(parts) >= 4:
                bwu, region, outlet, scene = parts[:4]
            else:
                bwu = region = outlet = scene = ""

            # Gather data
            data = [
                bwu, region, outlet, scene, self.bwu_var.get().replace('\n', '')
            ]
            defect_types = [
                "Switched OFF", "Screen/SAS", "Header not working", "Low visibility in header",
                "Shelf light", "Adjust shelves", "Top shelf", "Legal issue",
                "Visible content in header", "Short vertical insert", "Shelf light on comp", "Physical damage",
                "Header broken", "BWU not closing", "Broken flap", "Missing shelf",
                "Shelf strip base", "Shelf-strip insert", "Гнушка", "No POSM", "Client price tag over shelfstrip",
                "Header possible to install", "No content in Header", "EMPTY 1", "EMPTY 2", "EMPTY 3",
                "EMPTY 4", "EMPTY 5"
            ]
            defect_mapping = {
                "Switched\nOFF": "Switched OFF",
                "Screen/\nSAS": "Screen/SAS",
                "Header\nnot\nworking": "Header not working",
                "Low\nvisibility\nin header": "Low visibility in header",
                "Shelf\nlight": "Shelf light",
                "Adjust\nshelves": "Adjust shelves",
                "Top\nshelf": "Top shelf",
                "Legal\nissue": "Legal issue",
                "Visible\ncontent\nin header": "Visible content in header",
                "Short\nvertical\ninsert": "Short vertical insert",
                "Shelf light\non comp": "Shelf light on comp",
                "Physical\ndamage": "Physical damage",
                "Header\nbroken": "Header broken",
                "BWU not\nclosing": "BWU not closing",
                "Broken\nflap": "Broken flap",
                "Missing\nshelf": "Missing shelf",
                "Shelf\nstrip\nbase": "Shelf strip base",
                "Shelf-strip\ninsert": "Shelf-strip insert",
                "Гнушка": "Гнушка",
                "No POSM": "No POSM",
                "Client price\ntag over\nshelfstrip": "Client price tag over shelfstrip",
                "Header\npossible\nto install": "Header possible to install",
                "No content\nin header": "No content in Header",
                "EMPTY 1": "EMPTY 1",
                "EMPTY 2": "EMPTY 2",
                "EMPTY 3": "EMPTY 3",
                "EMPTY 4": "EMPTY 4",
                "EMPTY 5": "EMPTY 5"
            }
            for defect in defect_types:
                ui_defect = next((k for k, v in defect_mapping.items() if v == defect), None)
                data.append(defect if ui_defect and self.defect_vars.get(ui_defect, tk.BooleanVar(value=False)).get() else "")
            # Add comments
            comment = self.comments_entry.get().strip() if hasattr(self, 'comments_entry') else ""
            data.append(comment)

            # Write to Excel
            for col, value in enumerate(data, 1):
                ws[f"{get_column_letter(col)}{row}"] = value

            wb.save(self.excel_path)
            logging.debug(f"Saved catalog data for {image_name} to Excel")
        except Exception as e:
            logging.error(f"Error saving to Excel for {image_name}: {e}")
            messagebox.showerror("Error", f"Failed to save Excel data: {e}")

    def draw_defects_on_image(self, image_path):
        try:
            img = Image.open(image_path)
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", 30)
            except Exception:
                font = ImageFont.load_default()

            # Get selected defects, preserving original UI labels for EMPTY defects
            selected_defects = []
            for defect in self.defect_vars:
                if self.defect_vars[defect].get():
                    # Only replace newlines for non-EMPTY defects
                    if not defect.startswith("EMPTY"):
                        selected_defects.append(defect.replace('\n', ' '))
                    else:
                        selected_defects.append(defect)

            # Prepare text
            defect_text = "\n".join(selected_defects)
            if not defect_text:
                img.save(image_path, 'JPEG')
                return

            # Get image dimensions
            img_width, img_height = img.size
            padding = 10

            # Calculate text position (upper-right)

            # Estimate max text width for a 26-character string
            max_text_width = int(ImageFont.truetype("arial.ttf", 30).getbbox("A" * config.MAX_DEFECT_LENGTH)[2])  # Right coordinate of bbox gives width
            text_x = max(0, img_width - max_text_width - padding)  # Right-align with padding
            text_y = padding  # Start at top with padding


            # Draw yellow text
            draw.text((text_x, text_y), defect_text, fill="yellow", font=font)
            img.save(image_path, 'JPEG')
            logging.debug(f"Added defects to {image_path}")
        except Exception as e:
            logging.error(f"Error adding defects to {image_path}: {e}")

    def toggle_zoom(self, event, catalog_folder, images, catalog_window):
        if self.current_image_index >= len(images):
            return

        image_path = os.path.join(catalog_folder, images[self.current_image_index])
        try:
            img = Image.open(image_path)
            screen_width = catalog_window.winfo_screenwidth()
            screen_height = catalog_window.winfo_screenheight()

            # Cycle through zoom levels: 0 (thumbnail), 1 (full size with scroll wheel zoom)
            self.zoom_level = config.flip(self.zoom_level)

            if self.zoom_level == 0:
                # Thumbnail size
                max_size = (800, 600)
                img.thumbnail(max_size, Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                self.image_label.config(image=photo)
                self.image_label.image = photo  # Keep reference
            elif self.zoom_level == 1:
                # Full size with scroll wheel zoom in new window
                zoom_window = tk.Toplevel(catalog_window)
                zoom_window.title("Zoomed Image")
                # Set window size to fit within screen
                window_width = min(img.size[0], screen_width - 40)
                window_height = min(img.size[1], screen_height - 100)
                zoom_window.geometry(f"{window_width}x{window_height}+0+0")
                # Make window transient and grab focus
                zoom_window.transient(catalog_window)
                zoom_window.focus_set()
                # Close on Escape or click
                zoom_window.bind('<Escape>', lambda e: zoom_window.destroy())
                # Canvas for scrollable image
                canvas = tk.Canvas(zoom_window, width=window_width, height=window_height)
                canvas.pack(side="top", fill="both", expand=True)
                h_scrollbar = ttk.Scrollbar(zoom_window, orient="horizontal", command=canvas.xview)
                h_scrollbar.pack(side="bottom", fill="x")
                v_scrollbar = ttk.Scrollbar(zoom_window, orient="vertical", command=canvas.yview)
                v_scrollbar.pack(side="right", fill="y")
                canvas.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)
                # Initialize zoom scale
                self.current_zoom = 1.0
                # Store original image for zooming
                self.zoom_image = img
                # Image label
                photo = ImageTk.PhotoImage(img)
                image_label = tk.Label(canvas, image=photo)
                canvas.create_window((0, 0), window=image_label, anchor="nw")
                image_label.image = photo  # Keep reference
                # Update scroll region
                canvas.configure(scrollregion=(0, 0, img.size[0], img.size[1]))

                # Scroll wheel zoom
                def zoom(event):
                    logging.debug(f"Scroll event detected: delta={getattr(event, 'delta', 0)}")
                    # Adjust zoom level
                    if getattr(event, 'delta', 0) > 0 or event.num == 4:
                        self.current_zoom *= 1.1  # Zoom in
                    elif getattr(event, 'delta', 0) < 0 or event.num == 5:
                        self.current_zoom /= 1.1  # Zoom out
                    self.current_zoom = max(0.1, min(self.current_zoom, 5.0))  # Limit zoom range
                    # Resize image
                    new_width = int(img.size[0] * self.current_zoom)
                    new_height = int(img.size[1] * self.current_zoom)
                    zoomed_img = self.zoom_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(zoomed_img)
                    image_label.config(image=photo)
                    image_label.image = photo  # Keep reference
                    # Update scroll region
                    canvas.configure(scrollregion=(0, 0, new_width, new_height))

                # Bind scroll wheel to canvas and label (Windows and Linux)
                for widget in (canvas, image_label):
                    widget.bind("<MouseWheel>", zoom)  # Windows
                    widget.bind("<Button-4>", lambda e: zoom(type('event', (), {'num': 4})))  # Linux scroll up
                    widget.bind("<Button-5>", lambda e: zoom(type('event', (), {'num': 5})))  # Linux scroll down
                # Keyboard zoom fallback
                zoom_window.bind('<Control-plus>', lambda e: zoom(type('event', (), {'num': 4})))
                zoom_window.bind('<Control-minus>', lambda e: zoom(type('event', (), {'num': 5})))
                # Close window on click and reset zoom level
                image_label.bind("<Button-1>", lambda e: [zoom_window.destroy(), setattr(self, 'zoom_level', -1)])
        except Exception as e:
            logging.error(f"Error toggling zoom for {image_path}: {e}")

def start_gui():
    root = tk.Tk()
    app = ImageDownloaderGUI(root)
    root.mainloop()